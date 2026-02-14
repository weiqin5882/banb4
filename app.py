from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl.styles import PatternFill

from services.reconcile import ReconcileResult, reconcile

app = Flask(__name__)
LAST_RESULT: ReconcileResult | None = None


def read_excel_file(file_storage) -> pd.DataFrame:
    filename = (file_storage.filename or "").lower()
    if filename.endswith('.xls'):
        return pd.read_excel(file_storage, dtype=str, engine='xlrd')
    return pd.read_excel(file_storage, dtype=str)


def collect_manual(prefix: str) -> dict[str, str]:
    keys = ["order_no", "product_name", "status", "sales_amount", "cost_amount"]
    return {k: request.form.get(f"{prefix}_{k}", "").strip() for k in keys if request.form.get(f"{prefix}_{k}", "").strip()}


@app.get('/')
def index():
    return render_template('index.html')


@app.post('/api/reconcile')
def api_reconcile():
    global LAST_RESULT

    official_file = request.files.get('official_file')
    customer_file = request.files.get('customer_file')

    if not official_file or not customer_file:
        return jsonify({'error': '请同时上传官方订单表和客服统计表'}), 400

    try:
        official_df = read_excel_file(official_file)
        customer_df = read_excel_file(customer_file)

        result = reconcile(
            official_df,
            customer_df,
            official_manual=collect_manual('official'),
            customer_manual=collect_manual('customer'),
        )
        LAST_RESULT = result
        return jsonify(
            {
                'rows': result.rows,
                'summary': result.summary,
                'duplicates': result.duplicates,
                'mappings': result.mappings,
            }
        )
    except Exception as exc:  # noqa: BLE001
        return jsonify({'error': f'处理失败: {exc}'}), 400


@app.get('/api/export')
def export_excel():
    if LAST_RESULT is None:
        return jsonify({'error': '暂无可导出的比对结果，请先执行比对'}), 400

    output = BytesIO()
    df = pd.DataFrame(LAST_RESULT.rows)

    summary_df = pd.DataFrame([
        {'统计项': '总销售额', '数值': LAST_RESULT.summary['总销售额']},
        {'统计项': '总成本', '数值': LAST_RESULT.summary['总成本']},
        {'统计项': '总利润', '数值': LAST_RESULT.summary['总利润']},
        {'统计项': '订单总数', '数值': LAST_RESULT.summary['订单总数']},
    ])

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='对账结果', index=False)
        summary_df.to_excel(writer, sheet_name='汇总', index=False)

        ws = writer.book['对账结果']
        red_fill = PatternFill(fill_type='solid', fgColor='FFFFC7CE')
        for i, status in enumerate(df['状态标记'], start=2):
            if '亏损订单' in str(status):
                for cell in ws[i]:
                    cell.fill = red_fill

    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='对账结果.xlsx',
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)
